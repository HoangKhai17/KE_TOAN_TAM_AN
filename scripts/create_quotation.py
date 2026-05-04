"""
Tạo file Excel báo giá hợp đồng - Phần Mềm Quản Lý Nội Bộ Kế Toán Tâm An
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ─────────────────────────────────────────────
# Màu sắc & style dùng chung
# ─────────────────────────────────────────────
CLR_HEADER_DARK  = "1E3A5F"   # navy xanh đậm
CLR_HEADER_MED   = "2E75B6"   # xanh vừa
CLR_HEADER_LIGHT = "D6E4F0"   # xanh nhạt nền
CLR_ACCENT       = "E8F4FD"   # nền row xen kẽ
CLR_TOTAL        = "FFF2CC"   # vàng nhạt - tổng cộng
CLR_SAVE         = "E2EFDA"   # xanh lá nhạt - tiết kiệm
CLR_DANGER       = "FCE4D6"   # cam nhạt - chi phí cao
CLR_BREAKEVEN    = "FFFACD"   # vàng chanh - hoà vốn
CLR_WHITE        = "FFFFFF"

def ft(bold=False, size=11, color="000000", italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def border_thin():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def border_medium():
    s = Side(style="medium", color="888888")
    return Border(left=s, right=s, top=s, bottom=s)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

VND = '#,##0" đ"'
VND_M = '#,##0.0" tr"'

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def apply_row(ws, row, cells_styles):
    """cells_styles: list of (col, value, font, fill, alignment, border, number_format)"""
    for col, val, fnt, fll, aln, brd, nfmt in cells_styles:
        c = ws.cell(row=row, column=col, value=val)
        if fnt:  c.font      = fnt
        if fll:  c.fill      = fll
        if aln:  c.alignment = aln
        if brd:  c.border    = brd
        if nfmt: c.number_format = nfmt

def header_row(ws, row, cols_values, bg=CLR_HEADER_MED, fg="FFFFFF", size=11):
    for col, val in cols_values:
        c = ws.cell(row=row, column=col, value=val)
        c.font      = ft(bold=True, size=size, color=fg)
        c.fill      = fill(bg)
        c.alignment = align("center", "center", wrap=True)
        c.border    = border_thin()

def money_cell(ws, row, col, value, bg=None, bold=False):
    c = ws.cell(row=row, column=col, value=value)
    c.number_format = '#,##0'
    c.font      = ft(bold=bold, size=11)
    c.fill      = fill(bg) if bg else fill(CLR_WHITE)
    c.alignment = align("right", "center")
    c.border    = border_thin()

def text_cell(ws, row, col, value, bg=None, bold=False, italic=False, color="000000",
              h_align="left", wrap=False, size=11):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = ft(bold=bold, size=size, color=color, italic=italic)
    c.fill      = fill(bg) if bg else fill(CLR_WHITE)
    c.alignment = align(h_align, "center", wrap)
    c.border    = border_thin()


# ══════════════════════════════════════════════════════════════
#  SHEET 1 — BÁO GIÁ HỢP ĐỒNG
# ══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Báo Giá Hợp Đồng"
ws1.sheet_view.showGridLines = False

# Độ rộng cột
for col, w in [(1,4),(2,5),(3,30),(4,42),(5,18),(6,18)]:
    set_col_width(ws1, col, w)

ws1.row_dimensions[1].height = 14
ws1.row_dimensions[2].height = 38
ws1.row_dimensions[3].height = 18

# ── TIÊU ĐỀ ──────────────────────────────────────────────────
ws1.merge_cells("A2:F2")
c = ws1["A2"]
c.value     = "BÁO GIÁ PHẦN MỀM QUẢN LÝ NỘI BỘ"
c.font      = ft(bold=True, size=20, color=CLR_WHITE)
c.fill      = fill(CLR_HEADER_DARK)
c.alignment = align("center", "center")

ws1.merge_cells("A3:F3")
c = ws1["A3"]
c.value     = "Kế Toán Tâm An — Internal Management System"
c.font      = ft(italic=True, size=12, color="CCDDEE")
c.fill      = fill(CLR_HEADER_DARK)
c.alignment = align("center", "center")

# ── THÔNG TIN HỢP ĐỒNG ───────────────────────────────────────
info_rows = [
    ("Khách hàng",      "Công ty Kế Toán Tâm An"),
    ("Loại hợp đồng",   "Phát triển phần mềm — Bàn giao 1 lần (One-time Delivery)"),
    ("Ngày lập báo giá","04/05/2026"),
    ("Hiệu lực báo giá","30 ngày kể từ ngày lập"),
]
r = 5
ws1.merge_cells(f"A{r-1}:F{r-1}")
c = ws1.cell(row=r-1, column=1, value="THÔNG TIN HỢP ĐỒNG")
c.font      = ft(bold=True, size=11, color=CLR_WHITE)
c.fill      = fill(CLR_HEADER_MED)
c.alignment = align("left", "center")

for label, value in info_rows:
    bg = CLR_ACCENT if r % 2 == 0 else CLR_WHITE
    ws1.merge_cells(f"A{r}:B{r}")
    text_cell(ws1, r, 1, label, bg=bg, bold=True)
    ws1.merge_cells(f"C{r}:F{r}")
    text_cell(ws1, r, 3, value, bg=bg)
    r += 1

r += 1

# ── BẢNG CHI TIẾT GIÁ ────────────────────────────────────────
ws1.merge_cells(f"A{r}:F{r}")
c = ws1.cell(row=r, column=1, value="CHI TIẾT GÓI PHẦN MỀM")
c.font      = ft(bold=True, size=11, color=CLR_WHITE)
c.fill      = fill(CLR_HEADER_MED)
c.alignment = align("left", "center")
r += 1

header_row(ws1, r,
    [(1,"STT"),(2,"Module"),(3,"Tên hạng mục"),(4,"Mô tả chức năng chính"),(5,"Đơn giá (VND)"),(6,"Thành tiền (VND)")],
    bg=CLR_HEADER_DARK)
r += 1

modules = [
    # (stt, module, ten, mo_ta, gia)
    (1, "M1", "Hồ Sơ Khách Hàng",
     "Quản lý danh sách doanh nghiệp, hồ sơ công ty, lịch sử phân công nhân viên phụ trách",
     3_000_000),
    (2, "M2", "Quản Lý Nhân Sự",
     "Hồ sơ nhân viên, theo dõi tải công việc, phân quyền Admin / Nhân viên",
     2_500_000),
    (3, "M3a", "Quản Lý Công Việc — Core",
     "Tạo/giao/theo dõi task, luồng trạng thái 6 bước, nhật ký hoạt động, comment nội bộ",
     6_000_000),
    (4, "M3b", "Task Template 2 Lớp + 9 Chế Độ Lặp",
     "Task Type Library (Lớp 1) + Customer Task Schedule (Lớp 2) với 9 quy tắc lặp linh hoạt (ngày/tuần/tháng/quý/năm/tùy chỉnh)",
     5_000_000),
    (5, "M3c", "Tính Năng Nâng Cao Công Việc",
     "Subtask/Checklist, Task Dependencies, Time Tracking, Custom Fields theo loại công việc, Escalation tự động",
     4_500_000),
    (6, "M3d", "Nhắc Nhở & Thông Báo",
     "Cảnh báo deadline, email tổng hợp buổi sáng, escalation tự động khi task quá hạn",
     2_000_000),
    (7, "M4", "Báo Cáo & Thống Kê",
     "Dashboard KPI, ma trận báo cáo chéo (NV × KH × Loại), SLA Compliance, Aging, Velocity, Forecast, xuất Excel/PDF",
     4_500_000),
    (8, "M5", "Quản Lý Hồ Sơ & Giấy Tờ",
     "Tích hợp OneDrive (Microsoft Graph API), kho tài liệu theo KH, đính kèm file theo task, tìm kiếm",
     3_000_000),
    (9, "M6", "Cấu Hình Hệ Thống",
     "Quản lý tài khoản người dùng, danh mục hệ thống, cấu hình escalation, quản lý Customer Task Schedule",
     1_500_000),
    (10, "DEV", "Thiết Kế UI/UX & Frontend",
     "Thiết kế giao diện React.js, responsive, dashboard trực quan, calendar view, kanban board",
     3_000_000),
    (11, "DEV", "Hạ Tầng & Triển Khai",
     "Cấu hình VPS Vietnix, Docker Compose, Nginx + SSL, CI/CD cơ bản, hướng dẫn bàn giao vận hành",
     2_000_000),
]

total = sum(m[4] for m in modules)

for i, (stt, mod, ten, mo_ta, gia) in enumerate(modules):
    bg = CLR_ACCENT if i % 2 == 0 else CLR_WHITE
    text_cell(ws1, r, 1, stt,   bg=bg, h_align="center")
    text_cell(ws1, r, 2, mod,   bg=bg, bold=True, h_align="center")
    text_cell(ws1, r, 3, ten,   bg=bg, bold=True)
    text_cell(ws1, r, 4, mo_ta, bg=bg, wrap=True)
    ws1.row_dimensions[r].height = 42
    money_cell(ws1, r, 5, gia, bg=bg)
    money_cell(ws1, r, 6, gia, bg=bg)
    r += 1

# ── TỔNG CỘNG ────────────────────────────────────────────────
ws1.merge_cells(f"A{r}:D{r}")
c = ws1.cell(row=r, column=1, value="TỔNG CỘNG (trước chiết khấu)")
c.font      = ft(bold=True, size=12)
c.fill      = fill(CLR_TOTAL)
c.alignment = align("right", "center")
c.border    = border_medium()
money_cell(ws1, r, 5, total, bg=CLR_TOTAL, bold=True)
money_cell(ws1, r, 6, total, bg=CLR_TOTAL, bold=True)
ws1.cell(row=r, column=5).font = ft(bold=True, size=12)
ws1.cell(row=r, column=6).font = ft(bold=True, size=12)
ws1.cell(row=r, column=5).border = border_medium()
ws1.cell(row=r, column=6).border = border_medium()
r += 1

# Chiết khấu gói trọn bộ
final_price = 33_000_000
discount    = total - final_price   # 37,500,000 - 33,000,000 = 4,500,000
discount_pct = round(discount / total * 100)

ws1.merge_cells(f"A{r}:D{r}")
c = ws1.cell(row=r, column=1,
    value=f"Chiết khấu gói trọn bộ 6 Module ({discount_pct}%)  —  Mua toàn bộ hệ thống một lần")
c.font      = ft(bold=True, size=11, color="1A7C3E")
c.fill      = fill(CLR_SAVE)
c.alignment = align("right", "center")
c.border    = border_thin()

for col in [5, 6]:
    c2 = ws1.cell(row=r, column=col, value=-discount)
    c2.number_format = '#,##0'
    c2.font      = ft(bold=True, size=11, color="1A7C3E")
    c2.fill      = fill(CLR_SAVE)
    c2.alignment = align("right", "center")
    c2.border    = border_thin()
r += 1

ws1.merge_cells(f"A{r}:D{r}")
c = ws1.cell(row=r, column=1, value="GIÁ HỢP ĐỒNG CHÍNH THỨC")
c.font      = ft(bold=True, size=14, color=CLR_WHITE)
c.fill      = fill(CLR_HEADER_DARK)
c.alignment = align("right", "center")
c.border    = border_medium()
for col in [5, 6]:
    c2 = ws1.cell(row=r, column=col, value=final_price)
    c2.number_format = '#,##0'
    c2.font      = ft(bold=True, size=14, color="FFD700")
    c2.fill      = fill(CLR_HEADER_DARK)
    c2.alignment = align("right", "center")
    c2.border    = border_medium()
ws1.row_dimensions[r].height = 32
r += 2

# ── ĐIỀU KHOẢN ───────────────────────────────────────────────
ws1.merge_cells(f"A{r}:F{r}")
c = ws1.cell(row=r, column=1, value="ĐIỀU KHOẢN & PHƯƠNG THỨC THANH TOÁN")
c.font=ft(bold=True, size=11, color=CLR_WHITE); c.fill=fill(CLR_HEADER_MED); c.alignment=align("left","center")
r += 1

terms = [
    ("Loại hợp đồng",     "One-time Delivery — Phát triển & bàn giao 1 lần, không có phí duy trì định kỳ"),
    ("Đợt 1 — Ký hợp đồng", f"40% — {int(final_price * 0.4):,} đ"),
    ("Đợt 2 — Demo hoàn chỉnh", f"40% — {int(final_price * 0.4):,} đ"),
    ("Đợt 3 — Nghiệm thu & bàn giao", f"20% — {int(final_price * 0.2):,} đ"),
    ("Thời gian thực hiện", "60–90 ngày làm việc kể từ ngày ký hợp đồng"),
    ("Bảo hành miễn phí",  "30 ngày sau bàn giao (fix bug không phát sinh thêm phí)"),
    ("Mã nguồn",           "Bàn giao toàn bộ source code sau khi thanh toán đủ 100%"),
]
for label, value in terms:
    bg = CLR_ACCENT if r % 2 == 0 else CLR_WHITE
    ws1.merge_cells(f"A{r}:B{r}")
    text_cell(ws1, r, 1, label, bg=bg, bold=True)
    ws1.merge_cells(f"C{r}:F{r}")
    text_cell(ws1, r, 3, value, bg=bg, wrap=True)
    ws1.row_dimensions[r].height = 28
    r += 1

r += 1
ws1.merge_cells(f"A{r}:F{r}")
c = ws1.cell(row=r, column=1,
    value="* Chi phí vận hành hàng năm (VPS Vietnix ~4,800,000 đ/năm) do khách hàng tự chi trả, không thuộc phạm vi hợp đồng này.")
c.font=ft(italic=True, size=10, color="666666"); c.alignment=align("left","center",True)
ws1.row_dimensions[r].height = 24


# ══════════════════════════════════════════════════════════════
#  SHEET 2 — SO SÁNH CHI PHÍ THEO NĂM
# ══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("So Sánh Chi Phí Theo Năm")
ws2.sheet_view.showGridLines = False

for col, w in [(1,22),(2,18),(3,18),(4,18),(5,4),(6,18),(7,18),(8,18)]:
    set_col_width(ws2, col, w)

# Tiêu đề
ws2.merge_cells("A1:H1")
c = ws2["A1"]
c.value="SO SÁNH CHI PHÍ THEO NĂM: PHẦN MỀM TÂM AN vs CLICKUP"
c.font=ft(bold=True,size=16,color=CLR_WHITE); c.fill=fill(CLR_HEADER_DARK); c.alignment=align("center","center")
ws2.row_dimensions[1].height = 36

ws2.merge_cells("A2:H2")
c = ws2["A2"]
c.value = (
    "Giá hợp đồng 1 lần: 33,000,000 đ  |  "
    "Chi phí vận hành hàng năm: 4,800,000 đ/năm  |  "
    "ClickUp Business: $10/user/tháng ≈ 250,000 đ/user/tháng"
)
c.font=ft(italic=True,size=10,color="555555"); c.alignment=align("center","center")
ws2.row_dimensions[2].height = 22

# ── Bảng cho từng mức user ───────────────────────────────────
USD_RATE = 25_000  # 1 USD = 25,000 VND
CLICKUP_USD_PER_USER_MONTH = 10
SOFTWARE_ONETIME = 33_000_000
VPS_PER_YEAR = 4_800_000
YEARS = 5

def clickup_annual(n_users):
    return CLICKUP_USD_PER_USER_MONTH * n_users * 12 * USD_RATE

user_scenarios = [
    (5,  "5 người dùng"),
    (10, "10 người dùng"),
    (20, "20 người dùng"),
]

r = 4
for n_users, label in user_scenarios:
    cu_annual = clickup_annual(n_users)

    # Header kịch bản
    ws2.merge_cells(f"A{r}:H{r}")
    c = ws2.cell(row=r, column=1,
        value=f"KỊCH BẢN: {label}  |  ClickUp = {cu_annual//1_000_000:,.0f} tr/năm  |  Giá trị so sánh mỗi năm")
    c.font=ft(bold=True,size=12,color=CLR_WHITE); c.fill=fill(CLR_HEADER_MED); c.alignment=align("left","center")
    ws2.row_dimensions[r].height = 26
    r += 1

    # Column headers
    header_row(ws2, r, [
        (1,"Năm"),
        (2,"Phần Mềm Tâm An\n(Chi phí năm đó)"),
        (3,"Phần Mềm Tâm An\n(Tổng cộng lũy kế)"),
        (4,"ClickUp\n(Chi phí năm đó)"),
        (5,""),
        (6,"ClickUp\n(Tổng cộng lũy kế)"),
        (7,"Chênh lệch lũy kế\n(ClickUp − Tâm An)"),
        (8,"Nhận xét"),
    ], bg=CLR_HEADER_DARK)
    r += 1

    taman_cum = 0
    clickup_cum = 0
    for yr in range(1, YEARS + 1):
        taman_year = SOFTWARE_ONETIME if yr == 1 else VPS_PER_YEAR
        taman_cum += taman_year
        clickup_cum += cu_annual
        diff = clickup_cum - taman_cum

        if yr == 1:
            remark = "Năm đầu: đầu tư ban đầu cao hơn"
            bg_remark = CLR_DANGER
        elif abs(diff) < cu_annual * 0.3:
            remark = "⚖️ Tiệm cận điểm hoà vốn"
            bg_remark = CLR_BREAKEVEN
        elif diff > 0:
            remark = f"✅ Tiết kiệm lũy kế {diff/1_000_000:.1f} tr so với ClickUp"
            bg_remark = CLR_SAVE
        else:
            remark = f"ClickUp vẫn rẻ hơn {abs(diff)/1_000_000:.1f} tr"
            bg_remark = CLR_DANGER

        row_bg = CLR_ACCENT if yr % 2 == 0 else CLR_WHITE

        text_cell(ws2, r, 1, f"Năm {yr}", bg=row_bg, bold=True, h_align="center")
        money_cell(ws2, r, 2, taman_year, bg=row_bg)
        money_cell(ws2, r, 3, taman_cum,  bg=row_bg, bold=True)
        money_cell(ws2, r, 4, cu_annual,  bg=row_bg)
        text_cell(ws2, r, 5, "vs", bg=row_bg, h_align="center", color="888888")
        money_cell(ws2, r, 6, clickup_cum, bg=row_bg, bold=True)

        c_diff = ws2.cell(row=r, column=7, value=diff)
        c_diff.number_format = '#,##0'
        c_diff.font = ft(bold=True, color="1A7C3E" if diff >= 0 else "CC0000")
        c_diff.fill = fill(CLR_SAVE if diff >= 0 else CLR_DANGER)
        c_diff.alignment = align("right", "center")
        c_diff.border = border_thin()

        text_cell(ws2, r, 8, remark, bg=bg_remark, wrap=True, size=10)
        ws2.row_dimensions[r].height = 24
        r += 1

    r += 2  # khoảng cách giữa các kịch bản


# ── Phần tổng kết ─────────────────────────────────────────────
ws2.merge_cells(f"A{r}:H{r}")
c = ws2.cell(row=r, column=1, value="TÓM TẮT & KẾT LUẬN")
c.font=ft(bold=True,size=12,color=CLR_WHITE); c.fill=fill(CLR_HEADER_DARK); c.alignment=align("left","center")
ws2.row_dimensions[r].height = 26
r += 1

conclusions = [
    ("5 users",  "Hoà vốn năm 3. Từ năm 4 trở đi tiết kiệm 7.2 tr/năm so với ClickUp. Sau 5 năm tiết kiệm ~14 tr."),
    ("10 users", "Hoà vốn trước năm 2. Sau 5 năm tiết kiệm ~90 tr so với ClickUp (giá trị rất lớn)."),
    ("20 users", "Hoà vốn trong năm 1. Sau 5 năm tiết kiệm ~195 tr. ROI cực cao khi đội nhóm lớn."),
    ("Lợi thế khác",
     "Không giới hạn số user — thêm nhân viên mới không tốn thêm phí. "
     "Dữ liệu hoàn toàn thuộc sở hữu khách hàng. "
     "Tuỳ chỉnh theo đúng nghiệp vụ kế toán — ClickUp là công cụ đa ngành, không tối ưu cho quy trình kế toán."),
]
for label, text in conclusions:
    bg = CLR_ACCENT if r % 2 == 0 else CLR_WHITE
    ws2.merge_cells(f"A{r}:B{r}")
    text_cell(ws2, r, 1, label, bg=bg, bold=True)
    ws2.merge_cells(f"C{r}:H{r}")
    text_cell(ws2, r, 3, text, bg=bg, wrap=True, size=10)
    ws2.row_dimensions[r].height = 36
    r += 1


# ══════════════════════════════════════════════════════════════
#  LƯU FILE
# ══════════════════════════════════════════════════════════════
out_path = r"d:\WorkSpace_ADA_Bamboo\PROJECT_BBOTECH\KE_TOAN_TAM_AN\docs\BaoGia_PhanMem_KeToanTamAn_v2.xlsx"
wb.save(out_path)
print(f"[OK] Da tao file: {out_path}")
